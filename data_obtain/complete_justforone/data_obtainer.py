#!/usr/bin/python3
#################################################################################################################
#      This script obtain informationd desired (for more/less information this script should be modified by hand) 
#       from a Zacros simulation. The aim of this script is to obtain all the desired information for a specific 
#       simulation, so my work-way is to run simulations and use the obtain evolution by time to see all informa-
#       tion of each simulation ( this is authomatic and doesnt need any change); then use the first option 
#       modifying the general information class to obtain only the desired information.
#       For last use the serie data obtainer with superscripts to obtain all the desired information from many
#       simulations in differents folders.
#      
#
#################################################################################################################
import numpy as np
import scm.plams
import pyzacros as pz
import pandas as pd
import os
import re
import sys
import xlsxwriter
import openpyxl

# General clas that have importat information for the simulation
class data_C():
    # Colors for output in terminal
    Black= "\u001b[30;1m"
    Red= "\u001b[31;1m"
    Green= "\u001b[32;1m"
    Yellow= "\u001b[33;1m"
    Blue= "\u001b[34;1m"
    Magenta= "\u001b[35;1m"
    Cyan= "\u001b[36;1m"
    White= "\u001b[37;1m"
    END="\u001b[0m"
    # Data for general acces
    options=None
    # procstat object
    pstat=None
    # specnum object
    spec=None
    # net object
    net=None
    # total steps runned in this simulation
    total_steps=None
    # Actual path
    CWD_path=None
    # arguments passed to the script
    argv=None
    # Names of all surface species
    surf_specs_names=None
    # Names of all gas species
    gas_specs_names=None
    # gas molar fractions
    gas_molar_fracs=None
    # Elementar steps names (with out _fwd or _rev) 
    elementary_steps_names=None
    elementary_steps_names_full=None
    number_of_lattice_sites=None


# Class that ask the user about which type of information is desired. It contains the ask_now function
#   which ask you to select between manual (script shoul be changed in  general_information class) or automatic
#   (prints all information)
class asker(data_C):
    def __init__(self):
        super().__init__()
        # data zacros.job
        self.ask_now()


    def ask_now(self):
        print(self.Red+" Hi!! choose your options for data analysis"+self.END)
        strings = [("     Build a sequence of 0 and 1, for false and true\n"),
                   (self.Yellow + " a)"+self.END+" Obtain general information to xlsx (manual)\n"),
                   (self.Yellow + " b)"+self.END+" Obtain evolution of the mean time by process(auto)\n"),
                   (self.Yellow + " d)"+self.END+" Type exit to stop\n")]
        ask_user=True
        for i in strings:
            print(i)
        while (ask_user==True):
            options = input(" Give your sequence: ")
            if (options=="exit"):
                ask_user=False
                quit()
            elif (len(options)!=2):
                print(self.Red+ "\n    This is not a predetermined options to be runned!!\n"+ self.END)
            elif re.match('^[0-1]*$', options):
                print(self.Green + "Starting process"+self.END)
                data_C.options=options
                ask_user=False
            else:
                print(self.Red+ "\n    This is not a predetermined options to be runned!!\n"+ self.END)


# Using the pyZacros library this class build the results object that then will be used
class initial(data_C):
    def __init__(self,work):
        super().__init__()
        # Procees stat
        # Return the statistics from the ‘procstat_output.txt’
        #   file in a form of a list of dictionaries. Below is shown 
        #   an example of the procstat_output.txt for a zacros calculation.
        data_C.pstat = work.get_process_statistics()
        # Specnum datos
        # Return the provided quantities from the specnum_output.txt 
        #   file in a form of a dictionary. Below is shown an example
        #   of the specnum_output.txt for a zacros calculation.
        data_C.spec = work.provided_quantities()

        # Return the reactions from the ‘general_output.txt’ file.
        data_C.net = work.get_reaction_network()

        # return the number of total steps of the simulation
        data_C.total_steps = work.number_of_snapshots()
        
        data_C.number_of_lattice_sites= work.number_of_lattice_sites()
        
        # Obtain the list of all elementary steps take into accfount in the simulation
        data_C.elementary_steps_names_full=work.elementary_steps_names()
        
        temp_list=[]
        # Removing the "_fwd" and "_rev" suffix, and remove duplicates
        for item in data_C.elementary_steps_names_full:
            item2=item[:-4]
            if (not(item2 in temp_list)):
                temp_list.append(item2)
        data_C.elementary_steps_names=temp_list
        
        # Options from the asker:  Exit
        if (data_C.options=="000"):
            print(self.Red + "\n   hmmmmm... nothing to do... i see"+ self.END)
            quit()

        # Initialize method class
        methods = methods_C()
        methods.input_infor()
        # Options from the asker:
        if (data_C.options[0]=="1"):
            # xml file for general information
            methods.general_information()
        if (data_C.options[1]=="1"):
            # info of times and others
            all_info()
# This class contains differents methods that then will be used to treat the data
class methods_C(data_C):
    def __init__(self):
        super().__init__()
        self.method_list=["ocupacion",
                          "proc_dime",
                          "general_information",
                          "all_info"]
    # Obtaing the name of all surface species, the gas molar fractions and the gas species nammes 
    #   from the simulation_input file
    def input_infor(self):
        with open(data_C.CWD_path+"/simulation_input.dat") as data_file:
            for line in data_file:
                if line.startswith("surf_specs_names"):
                    data_C.surf_specs_names = (line.split()[1:])
                if line.startswith("gas_molar_fracs"):
                    data_C.gas_molar_fracs= line.split()[1:]
                if line.startswith("gas_specs_names"):
                    data_C.gas_specs_names= line.split()[1:]

    # Return the fractional ocupation on the lattice of the specie given
    #   Input:
    #       -spec:
    #           specnum object (see pyZacros manual for more information)
    #       -name:
    #           String with the name of the specie desired
    #       -surf_specs_names:
    #           List with all the species that can be in the surface
    #   Output:
    #       - return the ocupation from the total (empty site are not taken into account)    
    def ocupacion(self,name,surf_specs_names):
        # Decuelve el valor en porcentaje de ocupacion
        total = 0
        for spc in surf_specs_names:
            total += data_C.spec.get(spc)[-1]
        name_val = data_C.spec.get(name)[-1]
        if (total==0):
            return 0
        return name_val/total

    # Retur the total number of events 
    #   Input:
    #       -name:
    #           String with the name of the specie desired
    #       -num_events_proc:
    #           Object with al the numer of events at one time. This object is obtained from the pyZacros function 
    #               "pstat=result.get_process_statistics()" and the if you select for example last step "dats=pstat[-1]"
    #               and now you can obtain the num_events_proc object as "num_events_proc=(dats["number_of_events"])".
    #       -rev:
    #           If you want to reverse the reaction, take fwd as rev and rev as fwd.
    #   Output:
    #       - return the number of procces (it rest fwd and rev giving the real total number of process with direction.)
    def proc_dime(self, name, num_events_proc, rev=False):
        name_fwd = name + "_fwd"
        name_rev = name + "_rev"
        if (rev==False):
            av_t_1 = num_events_proc[name_fwd]
            av_t_2 = num_events_proc[name_rev]
            av_t_1 = av_t_1 - av_t_2
            return av_t_1 # devuelve el numero de procesos hacia delante de esta reacción
        elif (rev==True):
            av_t_1 = num_events_proc[name_fwd]
            av_t_2 = num_events_proc[name_rev]
            av_t_1 = av_t_2 - av_t_1
            return av_t_1 # devuelve el numero de procesos hacia delante de esta reacción
    
    # Manual data obtainer. 
    #   This method needs of modification for different simulations, but gives a general idea of
    #    how you should use this section; moverover, helps you to determine the information wanted
    #    before running the "data_obtainer_serie.py"
    def general_information(self):
        xlx_path = data_C.CWD_path+"/results.xlsx"
        if (not (os.path.exists(xlx_path))):
            print(self.Yellow + "\n         No xlx input file, building empty one\n"+self.END)
            wb = openpyxl.Workbook()
            ws =  wb.active
            ws.title = "Sheet1"
        else:
            wb = openpyxl.load_workbook(xlx_path)
            ws = wb['Sheet1']

        
        # Select the letter of the column
        number_col = 2
        letter = xlsxwriter.utility.xl_col_to_name(number_col)

        # Geting information from las frame
        dats=data_C.pstat[-1]

        # Obtaining the num_event_proc object from las step of the simulation
        num_events_proc=(dats["number_of_events"])

        # Now a dataframe with all the information to print to the xlsx file will be builded.
        column = pd.DataFrame({'Nueva':[
        str(data_C.spec.get("Temperature")[-1]),
        "time_simulation",
        dats["time"],
        "P_parc",
        float(data_C.gas_molar_fracs[0]), # PCO2
        float(data_C.gas_molar_fracs[2]), # PH2
        float(data_C.gas_molar_fracs[1]), # P
        float(data_C.gas_molar_fracs[3]), # P
        "GAS_prod",
        data_C.spec.get("CO2")[-1],
        data_C.spec.get("H2")[-1],
        data_C.spec.get("CO")[-1],
        data_C.spec.get("H2O")[-1],
        "ADS",
        data_C.spec.get("CO2*")[-1],
        data_C.spec.get("H2*")[-1],
        data_C.spec.get("CO*")[-1],
        data_C.spec.get("H2O*")[-1],
        "OCUPACION",
        self.ocupacion("CO2*", data_C.surf_specs_names),
        self.ocupacion("CO*", data_C.surf_specs_names),
        self.ocupacion("H2*", data_C.surf_specs_names),
        self.ocupacion("H2O*", data_C.surf_specs_names),
        self.ocupacion("O*", data_C.surf_specs_names),
        self.ocupacion("H*", data_C.surf_specs_names),
        self.ocupacion("HCOO*", data_C.surf_specs_names),
        self.ocupacion("traCOOH*", data_C.surf_specs_names),
        self.ocupacion("cisCOOH*", data_C.surf_specs_names),
        self.ocupacion("COH*", data_C.surf_specs_names),
        self.ocupacion("HCO*", data_C.surf_specs_names),
        self.ocupacion("OH*", data_C.surf_specs_names),
        "Num_evnt", # DE lso procesos primero pondremos las formacion de cooh-t y 
        self.proc_dime("cCOOH_dissociation_2",num_events_proc), #ruptura del CO2 y HCOO form que son las dos vias principales
        self.proc_dime("COH_formation", num_events_proc, True),
        self.proc_dime("CO2_dissociation", num_events_proc),
        self.proc_dime("HCO_formation", num_events_proc),
        "Num_event_sectCOOH",
        self.proc_dime("cCOOH_dissociation_1", num_events_proc), #cCOOH-> COH +O cCOOH_dissociation_1
        self.proc_dime("cCOOH_dissociation_2", num_events_proc), #cCOOH-> CO +OH cCOOH_dissociation_2
        "OC_Ads",
        self.proc_dime("CO_de-adsorption", num_events_proc)
        ]})
        # Printing to file
        for index, row in column.iterrows():
            cell = '{letter}{val}'.format(letter=letter, val=(index + 2 ))
            ws[cell] = row[0]
        
        wb.save('results.xlsx')

# This class (method class) gives many information of the simulation that is analyzed. Being this information
#       -> Gas molecules changes
#       -> Surface ocupation
#       -> Number of events (fwd and rev)
#       -> Number of events (mean)
#  This data is given with their temporal evolution.
class all_info(data_C):
    def __init__(self):
        super().__init__()
        # Building top directory to save information
        i=0
        while (os.path.exists("DataZacros%s" %(i))):
            i += 1
        os.mkdir("DataZacros%s" % (i))
        # Building xlx file
        self.xlx_path = data_C.CWD_path+ ("/DataZacros%s" % (i))+"/results.xlsx"
        if (not (os.path.exists(self.xlx_path))):
            print(self.Yellow + "\n         No xlx input file, building empty one\n"+self.END)
            self.wb = openpyxl.Workbook()
            ws =  self.wb.active
            ws.title = "Sheet1"
        self.inti_specnum()

    def inti_specnum(self):
        # Print informaiton related to specnum 
        self.print_spec_info("surface_info", data_C.surf_specs_names,data_C.spec)   

        # Now in a new sheet we will write the information of gas species
        self.print_spec_info("gas_proc",data_C.gas_specs_names,data_C.spec)

        # Now we will obtain the process information for all reaction
        ######   But we will obtain relative process resting the fwd and rev information 
        ######   returning the number of events
        self.print_pstat_info_full("proc_full_info",
                        "number_of_events",
                        data_C.elementary_steps_names_full,
                        data_C.pstat)
        self.print_pstat_info("proc_info",
                        "number_of_events",
                        data_C.elementary_steps_names,
                        data_C.pstat)
        self.wb.remove(self.wb["Sheet1"])
        self.wb.save(self.xlx_path)
        
    # Prints the information relative to the surface occupation
    def print_spec_info(self,sheet_name, list_spec, data_object):
        ws=self.wb.create_sheet(sheet_name)
        self.write_time(ws,"spec")
        number_col=2
        for specie in list_spec:
            letter = xlsxwriter.utility.xl_col_to_name(number_col)
            list_things=data_object[specie]
            # Building as dataframe
            column = pd.DataFrame({'Nueva':
                list_things
                })
            # Select the cell to write the columnheader
            cell = '{letter}{val}'.format(letter=letter, val=(1))
            ws[cell] = specie
            # Now writing information of thi cell
            if (sheet_name=="surface_info"):
                for index, row in column.iterrows():
                    cell = '{letter}{val}'.format(letter=letter, val=(index + 2 ))
                    ws[cell] = row[0]/data_C.number_of_lattice_sites
            else:
                for index, row in column.iterrows():
                    cell = '{letter}{val}'.format(letter=letter, val=(index + 2 ))
                    ws[cell] = row[0]
            number_col += 1

    # Print the number of events of each process (fwd and rev separated)
    def print_pstat_info_full(self,sheet_name,information, list_pstat, data_object):
        ws=self.wb.create_sheet(sheet_name)
        self.write_time(ws,"pstat")
        number_col=2
        # Loop ovr the list_pstat valuex [ex: process_fwd]
        for specie in list_pstat:
            list_things=[]
            for time_info in data_object:
                list_things.append(time_info[information][specie])
            letter = xlsxwriter.utility.xl_col_to_name(number_col)
            # Building as dataframe
            column = pd.DataFrame({'Nueva':
                list_things
                })
            # Select the cell to write the columnheader
            cell = '{letter}{val}'.format(letter=letter, val=(1))
            ws[cell] = specie
            # Now writing information of thi cell
            for index, row in column.iterrows():
                cell = '{letter}{val}'.format(letter=letter, val=(index + 2 ))
                ws[cell] = row[0]
            number_col += 1

    # Print the number of events using the difference value
    def print_pstat_info(self,sheet_name,information,list_pstat,data_object):
        ws=self.wb.create_sheet(sheet_name)
        self.write_time(ws,"pstat")
        number_col=2
        for specie in list_pstat:
            list_things=[]
            for i in range(len(data_C.pstat)):
                dats=data_object[i]
                num_events_proc=(dats[information])
                list_things.append(methods_C.proc_dime(self,specie,num_events_proc))
            letter = xlsxwriter.utility.xl_col_to_name(number_col)
            # Building as dataframe
            column = pd.DataFrame({'Nueva':
                list_things
                })
            # Select the cell to write the columnheader
            cell = '{letter}{val}'.format(letter=letter, val=(1))
            ws[cell] = specie
            # Now writing information of thi cell
            for index, row in column.iterrows():
                cell = '{letter}{val}'.format(letter=letter, val=(index + 2 ))
                ws[cell] = row[0]
            number_col += 1

    # Writes the time evolution  on the sercond column of each sheet of the results.xlsx
    def write_time(self,worksheet,type):
        # type spec or pstat
        if (type=="pstat"):
            time_list=[]
            for item in data_C.pstat:
                time_list.append(item["time"])
        elif(type=="spec"):
            time_list=data_C.spec["Time"]
        i=2
        letter = xlsxwriter.utility.xl_col_to_name(1)
        for item in time_list:
            cell='{letter}{val}'.format(letter=letter,val=(i))
            worksheet[cell] =  item
            i+=1

string = "set key autotitle columnhead"

if __name__ == "__main__":
    dir_path = os.getcwd()
    work = pz.ZacrosJob.load_external(path=dir_path,finalize=True)
    result = work.results
    work.status="successful"
    asker()
    data_C.CWD_path=dir_path
    data_C.argv = sys.argv[:]
    app = initial(result)
